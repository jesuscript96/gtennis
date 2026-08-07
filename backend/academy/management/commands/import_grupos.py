"""Importa los grupos de la pestaña «GRUPOS TODOS» (bloque DIVISIONES) del
calendario y crea las relaciones responsable–jugador (#2):

  * Entrenador PRINCIPAL (prioridad 1) = el de la propia columna/división.
  * Entrenadores SECUNDARIOS (prioridad 2) = los de las divisiones vecinas
    (división ±1), que es «las columnas de alrededor».

Además fija la división de cada jugador (la columna en la que aparece) y da de
alta los entrenadores que falten (BLAS, PATRICIO, SERGIO G., …).

Uso:
    python manage.py import_grupos --file "~/Downloads/CALENDARIO 2026.xlsx" --dry-run
    python manage.py import_grupos --file "~/Downloads/CALENDARIO 2026.xlsx"
"""
import os
import re
import unicodedata

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from academy.models import Division, Entrenador, Jugador, ResponsableJugador

SHEET = "GRUPOS TODOS "
# Columna (1-based) -> nivel de división.
COL_DIV = {3: 1, 4: 2, 5: 3, 7: 4, 8: 5, 10: 6, 11: 7, 13: 8}
# Entrenador(es) de cada columna (fila de cabecera de sub-entrenador).
COL_COACH = {
    3: "VICTOR R./SERGIO G.",
    4: "JAVI",
    5: "EMILIO",
    7: "BLAS",
    8: "MARIO/JORGE I./SALVA",
    10: "PATRICIO",
    11: "SANTI/NACHO",
    13: "ALVARO M.",
}
FILA_PRIMER_JUGADOR = 71
FILA_PRIMER_JUGADOR_COL13 = 70  # div 8 no tiene sub-cabecera

# Variantes de nombre del calendario que corresponden a un alumno del roster
# (clave = 2 primeros tokens en el Excel -> 2 primeros tokens en el roster).
ALIAS = {
    ("javi", "ballester"): ("javier", "ballester"),   # Javier Ballester Galarza
    ("natalia", "botea"): ("natalia", "ioana"),        # Natalia Ioana Botea
    ("nik", "guilin"): ("xu", "guilin"),               # Xu GUILIN (Yu) NIK
}


def _norm(s):
    s = str(s or "").strip()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    s = re.sub(r"[.\s]+$", "", s)
    s = re.sub(r"\s+", " ", s)
    return s


class Command(BaseCommand):
    help = "Importa grupos (responsables por división) desde el calendario."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True)
        parser.add_argument("--dry-run", action="store_true")

    # ---- entrenadores -----------------------------------------------------
    def _match_coach(self, token, cache):
        """Devuelve el Entrenador para un token de cabecera; lo crea si falta."""
        key = _norm(token).upper()
        if not key:
            return None
        if key in cache:
            return cache[key]
        # 1) exacto (normalizado, sin acentos, sin punto final)
        for e in Entrenador.objects.all():
            if _norm(e.nombre).upper() == key:
                cache[key] = e
                return e
        # 2) prefijo por palabra (NACHO -> NACHO C), sin colisiones tipo SERGIO
        cand = [
            e for e in Entrenador.objects.all()
            if _norm(e.nombre).upper().startswith(key + " ")
        ]
        if len(cand) == 1:
            cache[key] = cand[0]
            return cand[0]
        # 3) crear nuevo
        nombre = _norm(token).upper()
        e = Entrenador.objects.create(nombre=nombre)
        self._creados_coach.append(nombre)
        cache[key] = e
        return e

    def _coaches_de_columna(self, header, cache):
        out = []
        for tok in str(header).split("/"):
            e = self._match_coach(tok, cache)
            if e and e not in out:
                out.append(e)
        return out

    # ---- jugadores --------------------------------------------------------
    def _match_jugador(self, nombre, jindex):
        key = _norm(nombre).lower()
        toks = key.split()
        if len(toks) < 2:
            return None
        first2 = tuple(toks[:2])
        first2 = ALIAS.get(first2, first2)
        cand = jindex.get(first2, [])
        return cand[0] if len(cand) == 1 else None

    @transaction.atomic
    def handle(self, *args, **opts):
        import openpyxl

        path = os.path.expanduser(opts["file"])
        if not os.path.exists(path):
            raise CommandError(f"No existe: {path}")
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        sheet = SHEET if SHEET in wb.sheetnames else next(
            (s for s in wb.sheetnames if "GRUPOS" in s.upper()), None
        )
        ws = wb[sheet]
        rows = list(ws.iter_rows(min_row=1, max_row=90, values_only=True))

        def cell(r, c):  # 1-based
            return rows[r - 1][c - 1] if r - 1 < len(rows) else None

        # Índice de jugadores por (primer, segundo) token del nombre.
        jindex = {}
        for j in Jugador.objects.filter(activo=True):
            t = _norm(j.nombre).lower().split()
            if len(t) >= 2:
                jindex.setdefault(tuple(t[:2]), []).append(j)

        divisiones = {n: Division.objects.get_or_create(nivel=n)[0] for n in COL_DIV.values()}
        self._creados_coach = []
        cache = {}
        coaches_por_div = {COL_DIV[c]: self._coaches_de_columna(COL_COACH[c], cache) for c in COL_DIV}

        matched, unmatched = [], []
        # Recolecta jugadores por columna/división.
        por_div = {}
        for col, div in COL_DIV.items():
            start = FILA_PRIMER_JUGADOR_COL13 if col == 13 else FILA_PRIMER_JUGADOR
            nombres = []
            for r in range(start, 83):
                v = cell(r, col)
                if v is None or str(v).strip() in ("", "\xa0"):
                    continue
                nombres.append(str(v).strip())
            por_div[div] = nombres

        n_resp = 0
        for div, nombres in sorted(por_div.items()):
            principal = coaches_por_div.get(div, [])
            vecinos = coaches_por_div.get(div - 1, []) + coaches_por_div.get(div + 1, [])
            for nombre in nombres:
                j = self._match_jugador(nombre, jindex)
                if j is None:
                    unmatched.append((div, nombre))
                    continue
                matched.append((div, nombre, j.nombre))
                if not opts["dry_run"]:
                    j.division = divisiones[div]
                    j.save(update_fields=["division"])
                    for e in principal:
                        ResponsableJugador.objects.update_or_create(
                            jugador=j, entrenador=e,
                            defaults={"prioridad": 1},
                        )
                        n_resp += 1
                    for e in vecinos:
                        if e in principal:
                            continue
                        ResponsableJugador.objects.get_or_create(
                            jugador=j, entrenador=e,
                            defaults={"prioridad": 2},
                        )
                        n_resp += 1

        self.stdout.write("Entrenadores por división:")
        for div in sorted(coaches_por_div):
            nombres = ", ".join(e.nombre for e in coaches_por_div[div])
            self.stdout.write(f"  D{div}: {nombres}")
        self.stdout.write(self.style.SUCCESS(
            f"\nJugadores casados: {len(matched)} | sin casar: {len(unmatched)} | "
            f"entrenadores creados: {len(self._creados_coach)} ({', '.join(self._creados_coach) or '—'}) | "
            f"responsables: {n_resp}"
        ))
        if unmatched:
            self.stdout.write("Sin casar (revisar nombre):")
            for div, nombre in unmatched:
                self.stdout.write(f"  D{div}: {nombre}")
        if opts["dry_run"]:
            self.stdout.write(self.style.WARNING("DRY-RUN: no se ha escrito nada."))
            transaction.set_rollback(True)
