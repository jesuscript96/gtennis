"""Reconcile the Jugador roster against the academy's official client exports.

These `.xlsx` files come straight from the school's management software (the
"Resultados" sheet, one row per alumno) and are the source of truth for *who*
is in the academy. Unlike the old `import_excel` (which scraped the manual
`PLANTILLA MAÑANAS` grid), this command:

  * keys every player by `codigo_cliente` (stable external id),
  * upserts contact / birth / category data from the sheet,
  * and — by default — DELETES every active player not present in the exports,
    so the end state is exactly "todos y solo" the players in the files.

Curated fields the exports don't contain (division, entrenador_responsable,
rencillas, contratos, foto) are preserved on re-run.

Usage:
    python manage.py import_alumnos \
        --file ~/Downloads/out.xlsx --file "~/Downloads/out (1).xlsx"

    # preview without touching the DB
    python manage.py import_alumnos --file ... --dry-run

    # keep players not in the files instead of deleting them
    python manage.py import_alumnos --file ... --no-prune
"""
import os
import re
from datetime import date

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from academy.models import Jugador

# Column indexes in the "Resultados" sheet (0-based).
COL_CODIGO = 3
COL_ALUMNO = 4
COL_EMAIL = 5
COL_MOVIL = 6
COL_NACIMIENTO = 7
COL_MATRICULA = 8
COL_OBSERVACIONES = 10

PARTICULAS = {"de", "del", "la", "las", "los", "y", "da", "di", "van", "der"}
_EMPTY = {"", "\xa0", "none"}


def _s(value):
    """Normalise a cell to a stripped string ('' for blanks / nbsp)."""
    if value is None:
        return ""
    s = str(value).strip()
    return "" if s.lower() in _EMPTY else s


def _norm_word(word, first):
    """Title-case a single word if it is all-lower or all-upper; leave words
    that are already mixed-case untouched. Handles `(Yu)` style wrappers and
    keeps Spanish particles lowercased when not leading."""
    core = word.strip("()")
    wrapped = core != word
    if core and (core.islower() or core.isupper()):
        low = core.lower()
        core = low if (not first and low in PARTICULAS) else low.capitalize()
    return f"({core})" if wrapped else core


def clean_name(value):
    s = _s(value)
    s = re.sub(r"[\s.]+$", "", s)          # trailing dots / spaces
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return ""
    words = s.split(" ")
    return " ".join(_norm_word(w, i == 0) for i, w in enumerate(words))


def clean_email(value):
    s = _s(value)
    return s.lower() if "@" in s else ""


def parse_dob(value):
    s = _s(value)
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if not m:
        return None
    d, mo, y = (int(x) for x in m.groups())
    try:
        return date(y, mo, d)
    except ValueError:
        return None


def classify(matricula):
    m = _s(matricula).upper()
    if m.startswith("PROFESIONAL"):
        return Jugador.Categoria.PROFESIONAL
    if m.startswith("ALTO RENDIMIENTO") or m.startswith("ALTO"):
        return Jugador.Categoria.ALTO_RENDIMIENTO
    return ""


class Command(BaseCommand):
    help = "Reconcila los jugadores contra los Excel oficiales de alumnos."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file", action="append", dest="files", required=True,
            help="Ruta a un export .xlsx (repetible).",
        )
        parser.add_argument("--sheet", default="Resultados")
        parser.add_argument(
            "--no-prune", action="store_true",
            help="No borrar jugadores que no estén en los Excel.",
        )
        parser.add_argument(
            "--dry-run", action="store_true",
            help="No escribe en la BD; solo informa.",
        )

    def _parse_files(self, paths, sheet):
        import openpyxl

        alumnos = {}          # codigo_cliente -> dict
        for raw in paths:
            path = os.path.expanduser(raw)
            if not os.path.exists(path):
                raise CommandError(f"No existe el fichero: {path}")
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            if sheet not in wb.sheetnames:
                sheet = wb.sheetnames[0]
            ws = wb[sheet]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row is None or len(row) <= COL_ALUMNO:
                    continue
                nombre = clean_name(row[COL_ALUMNO])
                if not nombre:
                    continue
                codigo = row[COL_CODIGO]
                try:
                    codigo = int(codigo) if codigo not in (None, "") else None
                except (TypeError, ValueError):
                    codigo = None
                if codigo is None:
                    self.stderr.write(f"  · Sin código, ignorado: {nombre}")
                    continue
                if codigo in alumnos:
                    self.stderr.write(
                        f"  · Código duplicado {codigo} ({nombre}); me quedo el 1º."
                    )
                    continue
                alumnos[codigo] = {
                    "codigo_cliente": codigo,
                    "nombre": nombre,
                    "categoria": classify(row[COL_MATRICULA]),
                    "email": clean_email(row[COL_EMAIL]),
                    "telefono": _s(row[COL_MOVIL]),
                    "fecha_nacimiento": parse_dob(row[COL_NACIMIENTO]),
                    "notas": _s(row[COL_OBSERVACIONES])[:200],
                }
        return alumnos

    @transaction.atomic
    def handle(self, *args, **opts):
        alumnos = self._parse_files(opts["files"], opts["sheet"])
        if not alumnos:
            raise CommandError("No se ha extraído ningún alumno de los Excel.")

        creados = actualizados = 0
        for codigo, data in sorted(alumnos.items()):
            obj = Jugador.objects.filter(codigo_cliente=codigo).first()
            # Fallback: casa por nombre exacto un jugador sin código (legacy).
            if obj is None:
                obj = Jugador.objects.filter(
                    codigo_cliente__isnull=True, nombre__iexact=data["nombre"]
                ).first()
            if obj is None:
                if not opts["dry_run"]:
                    Jugador.objects.create(activo=True, **data)
                creados += 1
            else:
                for field, value in data.items():
                    setattr(obj, field, value)
                obj.activo = True
                if not opts["dry_run"]:
                    obj.save()
                actualizados += 1

        codigos = set(alumnos)
        sobrantes = Jugador.objects.exclude(codigo_cliente__in=codigos)
        n_sobrantes = sobrantes.count()
        borrados = 0
        if not opts["no_prune"]:
            for j in sobrantes:
                self.stdout.write(f"  − fuera: {j.nombre} (cod {j.codigo_cliente})")
            if not opts["dry_run"]:
                borrados, _ = sobrantes.delete()

        total = Jugador.objects.count() if not opts["dry_run"] else len(alumnos)
        self.stdout.write(self.style.SUCCESS(
            f"\nAlumnos en Excel: {len(alumnos)} | creados: {creados} | "
            f"actualizados: {actualizados} | fuera de los Excel: {n_sobrantes}"
            + (" (conservados, --no-prune)" if opts["no_prune"]
               else f" (eliminados; {borrados} filas en cascada)")
            + f"\nTotal jugadores en BD: {total}"
        ))
        if opts["dry_run"]:
            self.stdout.write(self.style.WARNING("DRY-RUN: no se ha escrito nada."))
            transaction.set_rollback(True)
